import os
import subprocess
from openpyxl import load_workbook

def convert_sheet_to_pdf(file_path, sheet_number):
    # Get the workbook
    wb = load_workbook(file_path)

    # Get the specified sheet
    sheet = wb.worksheets[sheet_number - 1]

    # Get the value in cell A1
    cell_value = sheet['A1'].value

    # Check if the cell value is negative
    if isinstance(cell_value, (int, float)) and cell_value < 0:
        # Convert the Excel sheet to PDF using SBCCmd
        output_file = os.path.splitext(file_path)[0] + f"_{sheet_number}.pdf"
        command = f'SBCCmd -d "{file_path}" -o "{output_file}" -xlssheetorderselect {sheet_number}'
        subprocess.run(command, shell=True)

def convert_excel_files_to_pdf(root_folder, extension=".xlsx"):
    # Recursively traverse the directory structure
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            if file.endswith(extension):
                file_path = os.path.join(root, file)

                # Get the workbook and loop through all sheets
                wb = load_workbook(file_path)
                for sheet_number, sheet in enumerate(wb.worksheets, start=1):
                    # Convert each sheet to PDF if it contains a negative value in A1
                    convert_sheet_to_pdf(file_path, sheet_number)

def main():
    root_folder = r"C:\Users\dylan\OneDrive\Desktop\test\Excel"
    convert_excel_files_to_pdf(root_folder)

if __name__ == "__main__":
    main()